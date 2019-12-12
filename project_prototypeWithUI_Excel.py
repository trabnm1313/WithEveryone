import sys
sys.path.insert(0, "package")
import time
import os
from tkinter import *
from openpyxl import load_workbook
workbook = load_workbook(filename="styles.xlsx")
workbook2 = load_workbook(filename="data.xlsx")
local_data = workbook2.active
local_sheet = workbook.active

def makeData():
    for row in range(2, len(local_data["A"])+1):
        datalist = []
        for data in local_data[row]:
            if "" != data.value and not type(None) == type(data.value):
                datalist.append(data.value)
        if datalist != [] and len(datalist)-1 == len(local_data["1"])-1:
            data_x.append(datalist[1:])
    print(*data_x)

def makeStyles():
    for row in range(1, len(local_sheet["A"])+1):
        datalist = []
        if type(local_sheet["A" + str(row)].value) != type(None):
            for col in range(2, len(local_sheet[str(row)])+1):
                if type(local_sheet[chr(64 + col) + str(row)].value) != type(None):
                    datalist.append(local_sheet[chr(64 + col) + str(row)].value)
            styles[local_sheet["A" + str(row)].value] = datalist

#Def function but idk why it's here
def callback(var):
    data_input.append(var)
    intvar.set(0)

def btnClicked():
    os.system('echo ' + "https://docs.google.com/forms/d/e/1FAIpQLSeB29jvNlxJ27pktRm8tHXaKxOW7Vl08FFyvTfvpWr2IRAKrQ/viewform" + "|clip")
    link['text'] = "Copied!"
    link.after(3000, lambda: link.config(text="Copy quiz link here."))

#Def ui variable
ui = Tk()
ui.config(bg='pink')
ui.title("Personality Analysis(EIEI)")
ui.resizable()
ui.minsize(475, 100)
ui.maxsize(475, 100)

#Samples data from database
styles = {}
data_x = []

#Variable
data_input = []
buttons = dict()
intvar = IntVar()
text = Label(ui, text="Welcome to personal preference analysis program!.", anchor=CENTER, font=("AngsanaUPC", 20), bg='pink')
text2 = Label(ui, text="", anchor=CENTER, font=("AngsanaUPC", 20), bg='pink')
text.grid(row=0, column=0, columnspan=8, padx=80)
link = Button(ui, text="Copy quiz link here.", command=btnClicked, bd=2)
link.grid(row=3, column=3, sticky=S, pady=5)

#Def function and call function
def choice(var):
    count = 0
    buttons = {}
    texts = {}
    text['text'] = "What " + var + " style would you prefer?"
    for index in range(1, len(styles[var])+1):
        buttons[index] = Button(ui, text=index, command=lambda style=styles[var][index-1]: callback(style), width=5, height=2)
        buttons[index].grid(row=index+1, column=0, pady=2)
    for index in range(1, len(styles[var])+1):
        texts[index] = Label(ui, text=styles[var][index-1], font=("AngsanaUPC", 18), bg='pink')
        texts[index].grid(row=index+1, column=1, pady=2)
    ui.maxsize(ui.winfo_reqwidth(), len(buttons)*70 + (len(buttons)*20))
    ui.minsize(ui.winfo_reqwidth(), len(buttons)*70 + (len(buttons)*20))
    ui.wait_variable(intvar)
    for data in data_x:
        brk = 0
        for char in range(len(data_input)):
            if data[char] != data_input[char]:
                brk = 1
        if brk == 0:
            count += 1
    text2['text'] = "There is " + str(count) + " match you choice out of " + str(len(data_x))

    for button in buttons:
        buttons[button].grid_forget()
    for txt in texts:
        texts[txt].grid_forget()

def main():
    makeData()
    makeStyles()
    ppl = 0
    keep = -1
    startbtn = Button(ui, text="Click here to start", command=lambda: intvar.set(0), bd=2)
    startbtn.grid(row=2, column=3)
    ui.wait_variable(intvar)
    ui.maxsize(1000, 1000)
    text2.grid(row=1, column=0, columnspan=8, padx=80)
    startbtn.grid_forget()
    link.grid_forget()

    for style in styles:
        choice(style)
    ui.minsize(500, 300)
    ui.maxsize(500, 300)

    ava = int((data_x.count(data_input)/len(data_x))*100)

    if data_x.count(data_input) > 0:
        text2['text'] = "There is " + str(data_x.count(data_input)) + " people match you styles out of " + str(len(data_x)) + "!"
        Label(text="You're " + str(ava) + "% from all kind of people.", font=("AngsanaUPC", 20), bg='pink').grid(row=4, column=3)
    else:
        text2['text'] = "Sorry, You styles is not match any people in our data."


    text['text'] = "Thanks for making surveys."

#Call main() and UI.mainloop()
main()
ui.mainloop()
